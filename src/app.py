
import streamlit as st
import pandas as pd
import openpyxl
import os
import json
import uuid
import base64
import zipfile
import sys
import re
import subprocess
import tempfile
import logging
from typing import Optional, Tuple
from dotenv import load_dotenv

# Configure logging for debugging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Function to get the current git commit hash
def get_git_commit():
    try:
        return subprocess.check_output(['git', 'rev-parse', '--short', 'HEAD']).decode('ascii').strip()
    except:
        return "unknown"
    
def get_git_commit_date():
    try:
        return subprocess.check_output(['git', 'show', '-s', '--format=%ct', get_git_commit()]).decode('ascii').strip()
    except Exception:
        return "unknown"

# Load environment variables
load_dotenv()

# Set the metadata filepath environment variable to None initially
# This will prevent the form_generator module from trying to load a non-existent file on import
os.environ['METADATA_FILEPATH'] = ''

# Import the existing form generation functions
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.form_generator import (
    generate_form,
    generate_translation_file,
    initialize_option_sets
)

# Load the configuration settings from config.json
def load_config():
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)

            # Ensure settings section exists
            if "settings" not in config:
                config["settings"] = get_default_app_settings()

            return config
    except Exception as e:
        st.error(f"Error loading configuration: {str(e)}")
        return {
            "columns": get_default_column_mappings(),
            "settings": get_default_app_settings()
        }

# Save configuration to config.json
def save_config(config):
    """
    Save configuration to config.json
    """
    try:
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, indent=4, fp=f)
        return True
    except Exception as e:
        st.error(f"Error saving configuration: {str(e)}")
        return False

# Function to get default column mappings
def get_default_column_mappings():
    """
    Return the default column mappings
    """
    return {
        "QUESTION_COLUMN": "Question",
        "LABEL_COLUMN": "Label if different",
        "QUESTION_ID_COLUMN": "Question ID",
        "EXTERNAL_ID_COLUMN": "External ID",
        "DATATYPE_COLUMN": "Datatype",
        "VALIDATION_COLUMN": "Validation (format)",
        "MANDATORY_COLUMN": "Mandatory",
        "RENDERING_COLUMN": "Rendering",
        "LOWER_LIMIT_COLUMN": "Lower limit",
        "UPPER_LIMIT_COLUMN": "Upper limit",
        "DEFAULT_VALUE_COLUMN": "Default value",
        "CALCULATION_COLUMN": "Calculation",
        "SKIP_LOGIC_COLUMN": "Skip logic",
        "PAGE_COLUMN": "Page",
        "SECTION_COLUMN": "Section",
        "OPTION_SET_COLUMN": "OptionSet name",
        "TOOLTIP_COLUMN_NAME": "Tooltip",
        "TRANSLATION_SECTION_COLUMN": "Translation - Section",
        "TRANSLATION_QUESTION_COLUMN": "Translation - Question",
        "TRANSLATION_TOOLTIP_COLUMN": "Translation - Tooltip",
        "TRANSLATION_ANSWER_COLUMN": "Translation"
    }

def get_default_app_settings():
    """
    Return the default application settings
    """
    return {
        "SHEET_FILTER_PREFIX": "F\\d{2}"  # Default: "F" followed by 2 digits
    }

def get_download_link(filepath, filename):
    """
    Generate a download link for a file
    """
    with open(filepath, 'rb') as f:
        bytes = f.read()
        b64 = base64.b64encode(bytes).decode()
        href = f'<a href="data:file/json;base64,{b64}" download="{filename}">Download {filename}</a>'
        return href

def safe_file_handler(uploaded_file) -> Tuple[Optional[str], str]:
    """
    Safely handle uploaded file with proper error handling for Streamlit Cloud.
    
    Args:
        uploaded_file: Streamlit uploaded file object
        
    Returns:
        Tuple[Optional[str], str]: (temp_file_path, error_message)
    """
    try:
        # Check file size (limit to 50MB for safety)
        file_size = len(uploaded_file.getvalue())
        logger.info(f"Processing file: {uploaded_file.name}, size: {file_size / (1024*1024):.2f} MB")
        
        if file_size > 50 * 1024 * 1024:  # 50MB limit
            return None, "File size exceeds 50MB limit. Please use a smaller file."
        
        # Use tempfile for cross-platform compatibility
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix='formgen_') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            temp_path = tmp_file.name
            
        # Validate file immediately after creation
        try:
            wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
            try:
                sheet_names = wb.sheetnames
                logger.info(f"Successfully validated Excel file with {len(wb.sheetnames)} sheets")
            finally:
                wb.close()
        except Exception as e:
            # Clean up temp file if validation fails
            try:
                os.unlink(temp_path)
            except:
                pass
            return None, f"Invalid Excel file: {str(e)}"
            
        return temp_path, ""
        
    except MemoryError:
        return None, "File too large to process. Please reduce file size or split into smaller files."
    except Exception as e:
        logger.error(f"Error handling file: {str(e)}")
        return None, f"Error processing file: {str(e)}"

def cleanup_temp_file(file_path: Optional[str]) -> None:
    """
    Safely cleanup temporary files.
    
    Args:
        file_path: Path to temporary file to cleanup
    """
    if file_path and os.path.exists(file_path):
        try:
            os.unlink(file_path)
            logger.info(f"Cleaned up temporary file: {file_path}")
        except Exception as e:
            logger.warning(f"Could not clean up temp file {file_path}: {str(e)}")

def generate_forms_from_sheets(metadata_file, selected_sheets):
    """
    Generate forms from selected sheets
    """
    try:
        # Enhanced file validation with better error reporting
        try:
            logger.info(f"Validating Excel file: {metadata_file}")
            wb = openpyxl.load_workbook(metadata_file, read_only=True, data_only=True)
            sheet_count = len(wb.sheetnames)
            wb.close()
            logger.info(f"Excel file validated successfully with {sheet_count} sheets")
        except zipfile.BadZipFile:
            st.error("❌ Invalid Excel file format - file appears to be corrupted or not a valid Excel file.")
            st.info("💡 **Solutions:**")
            st.info("• Save your file as 'Excel Workbook (.xlsx)' format")
            st.info("• Try opening and re-saving the file in Excel")
            st.info("• Ensure the file wasn't corrupted during upload")
            return []
        except FileNotFoundError:
            st.error("❌ File not found - temporary file may have been cleaned up.")
            st.info("💡 Please try uploading the file again.")
            return []
        except MemoryError:
            st.error("❌ File too large to process - insufficient memory.")
            st.info("💡 **Solutions:**")
            st.info("• Try reducing the file size by removing unnecessary data")
            st.info("• Split large files into smaller ones")
            st.info("• Remove unused sheets from the Excel file")
            return []
        except PermissionError:
            st.error("❌ Permission denied - cannot access the file.")
            st.info("💡 Please try uploading the file again.")
            return []
        except Exception as e:
            st.error(f"❌ Excel file validation failed: {str(e)}")
            st.info("💡 **Troubleshooting:**")
            st.info("• Ensure file is in .xlsx, .xlsm, .xltx, or .xltm format")
            st.info("• Check that the file is not password protected")
            st.info("• Try saving with Excel compatibility mode if created in newer Excel versions")
            logger.error(f"Excel validation failed: {str(e)}")
            return []

        # Set the metadata filepath environment variable for the form_generator module
        os.environ['METADATA_FILEPATH'] = metadata_file
        logger.info(f"Set METADATA_FILEPATH to: {metadata_file}")

        # Initialize option sets with enhanced error handling
        if not st.session_state.option_sets_initialized:
            try:
                logger.info("Initializing option sets...")
                initialize_option_sets(metadata_file)
                st.session_state.option_sets_initialized = True
                logger.info("Option sets initialized successfully")
            except Exception as e:
                st.error(f"❌ Failed to initialize option sets: {str(e)}")
                st.info("💡 Please ensure your Excel file has an 'OptionSets' sheet with the correct format.")
                logger.error(f"Option sets initialization failed: {str(e)}")
                return []

        # Load current configuration to ensure it's used
        config = load_config()

        # Update form_generator module's configuration variables with current settings
        import src.form_generator as fg
        for key, value in config.get("columns", {}).items():
            if hasattr(fg, key):
                setattr(fg, key, value)

        generated_forms = []

        for sheet in selected_sheets:
            translations_data = {}
            try:
                # Start timing the generation
                import time
                start_time = time.time()

                # Generate form and translations with enhanced error handling
                logger.info(f"Generating form for sheet: {sheet}")
                try:
                    form, _, total_questions, total_answers, missing_option_sets = generate_form(sheet, translations_data, metadata_file)
                    translations = generate_translation_file(sheet, 'ar', translations_data)
                    logger.info(f"Successfully generated form for {sheet} - {total_questions} questions, {total_answers} answers")
                except MemoryError:
                    st.error(f"❌ Insufficient memory to process sheet '{sheet}'. Try reducing the sheet size.")
                    continue
                except KeyError as e:
                    st.error(f"❌ Missing required column in sheet '{sheet}': {str(e)}")
                    st.info("💡 Please check your column mappings in the Configuration page.")
                    continue
                except Exception as e:
                    st.error(f"❌ Error generating form for sheet '{sheet}': {str(e)}")
                    logger.error(f"Form generation failed for {sheet}: {str(e)}")
                    continue

                # End timing
                generation_time = time.time() - start_time

                # Use temp directory or ensure output directory exists and is writable
                try:
                    output_dir = 'generated_forms'
                    os.makedirs(output_dir, exist_ok=True)
                    # Test write permissions
                    test_file = os.path.join(output_dir, '.test_write')
                    with open(test_file, 'w') as f:
                        f.write('test')
                    os.remove(test_file)
                    logger.info(f"Output directory ready: {output_dir}")
                except (PermissionError, OSError) as e:
                    # Fallback to temp directory if can't write to generated_forms
                    output_dir = tempfile.mkdtemp(prefix='formgen_output_')
                    logger.warning(f"Using temporary output directory: {output_dir}")
                    st.warning(f"⚠️ Using temporary directory for outputs: {os.path.basename(output_dir)}")

                # Generate filenames
                form_filename = f"{sheet.replace(' ', '_')}.json"
                translation_filename = f"{sheet.replace(' ', '_')}_translations_ar.json"

                form_path = os.path.join(output_dir, form_filename)
                translation_path = os.path.join(output_dir, translation_filename)

                # Convert to JSON strings
                form_json = json.dumps(form, indent=2)
                translation_json = json.dumps(translations, indent=2, ensure_ascii=False)

                # Calculate file sizes
                form_size = len(form_json.encode('utf-8'))
                translation_size = len(translation_json.encode('utf-8'))

                # Count sections and pages
                num_pages = len(form.get('pages', []))
                num_sections = sum(len(page.get('sections', [])) for page in form.get('pages', []))

                # Save files
                with open(form_path, 'w', encoding='utf-8') as f:
                    f.write(form_json)

                with open(translation_path, 'w', encoding='utf-8') as f:
                    f.write(translation_json)

                generated_forms.append({
                    'sheet': sheet,
                    'form_path': form_path,
                    'translation_path': translation_path,
                    'total_questions': total_questions,
                    'total_answers': total_answers,
                    'form_json': form_json,
                    'translation_json': translation_json,
                    'generation_time': generation_time,
                    'form_size': form_size,
                    'translation_size': translation_size,
                    'num_pages': num_pages,
                    'num_sections': num_sections,
                    'missing_option_sets': missing_option_sets
                })

                st.success(f"Successfully generated form for {sheet}")

            except MemoryError:
                st.error(f"❌ Out of memory while processing sheet '{sheet}'. Please reduce file size.")
                logger.error(f"Memory error processing sheet {sheet}")
                st.info(f"⏭️ Skipping sheet '{sheet}' and continuing with the next one.")
            except Exception as e:
                st.error(f"❌ Error generating form for sheet '{sheet}': {str(e)}")
                logger.error(f"Error processing sheet {sheet}: {str(e)}")
                st.info(f"⏭️ Skipping sheet '{sheet}' and continuing with the next one.")

        return generated_forms
    except MemoryError:
        st.error("❌ Out of memory error. The file is too large to process.")
        st.info("💡 **Solutions:**")
        st.info("• Reduce the file size by removing unnecessary data")
        st.info("• Process fewer sheets at once")
        st.info("• Split large files into smaller ones")
        logger.error("Memory error during form generation")
        return []
    except Exception as e:
        st.error(f"❌ An unexpected error occurred: {str(e)}")
        logger.error(f"Unexpected error during form generation: {str(e)}")
        st.info("💡 Please check the file format and try again. If the issue persists, try with a smaller file.")
        return []

def show_configuration_page():
    st.title("🔧 Column Mapping Configuration")
    st.markdown("""
    ### Customize Excel Column Mappings

    Modify the column names used by the form generator to match your Excel file structure.
    """)

    # Load current configuration
    config = load_config()
    column_mappings = config.get("columns", get_default_column_mappings())
    app_settings = config.get("settings", get_default_app_settings())

    # Check if we have saved settings in localStorage
    if 'column_mappings' in st.session_state:
        column_mappings = st.session_state.column_mappings
    if 'app_settings' in st.session_state:
        app_settings = st.session_state.app_settings

    # Add option to upload a config file
    st.subheader("Import Configuration")
    uploaded_config = st.file_uploader(
        "Upload existing config.json file",
        type=['json'],
        help="Upload a previously saved config.json file to restore your settings"
    )

    if uploaded_config is not None:
        try:
            imported_config = json.load(uploaded_config)
            if "columns" in imported_config:
                column_mappings = imported_config["columns"]
                st.session_state.column_mappings = column_mappings
            if "settings" in imported_config:
                app_settings = imported_config["settings"]
                st.session_state.app_settings = app_settings
            st.success("Configuration imported successfully!")
        except Exception as e:
            st.error(f"Error importing configuration: {str(e)}")

    # Create tabs for different configuration sections
    tab1, tab2 = st.tabs(["Column Mappings", "Application Settings"])

    with tab1:
        # Create a form for the column configuration
        with st.form("column_mapping_form"):
            # Group related fields
            st.subheader("Basic Question Fields")
            col1, col2 = st.columns(2)
            with col1:
                column_mappings["QUESTION_COLUMN"] = st.text_input("Question Column", column_mappings.get("QUESTION_COLUMN", "Question"))
                column_mappings["LABEL_COLUMN"] = st.text_input("Label Column", column_mappings.get("LABEL_COLUMN", "Label if different"))
                column_mappings["QUESTION_ID_COLUMN"] = st.text_input("Question ID Column", column_mappings.get("QUESTION_ID_COLUMN", "Question ID"))
                column_mappings["EXTERNAL_ID_COLUMN"] = st.text_input("External ID Column", column_mappings.get("EXTERNAL_ID_COLUMN", "External ID"))

            with col2:
                column_mappings["DATATYPE_COLUMN"] = st.text_input("Datatype Column", column_mappings.get("DATATYPE_COLUMN", "Datatype"))
                column_mappings["MANDATORY_COLUMN"] = st.text_input("Mandatory Column", column_mappings.get("MANDATORY_COLUMN", "Mandatory"))
                column_mappings["RENDERING_COLUMN"] = st.text_input("Rendering Column", column_mappings.get("RENDERING_COLUMN", "Rendering"))
                column_mappings["TOOLTIP_COLUMN_NAME"] = st.text_input("Tooltip Column", column_mappings.get("TOOLTIP_COLUMN_NAME", "Tooltip"))

            st.subheader("Layout and Organization")
            col1, col2 = st.columns(2)
            with col1:
                column_mappings["PAGE_COLUMN"] = st.text_input("Page Column", column_mappings.get("PAGE_COLUMN", "Page"))
                column_mappings["SECTION_COLUMN"] = st.text_input("Section Column", column_mappings.get("SECTION_COLUMN", "Section"))
                column_mappings["OPTION_SET_COLUMN"] = st.text_input("Option Set Column", column_mappings.get("OPTION_SET_COLUMN", "OptionSet name"))

            st.subheader("Validation and Calculation")
            col1, col2 = st.columns(2)
            with col1:
                column_mappings["VALIDATION_COLUMN"] = st.text_input("Validation Column", column_mappings.get("VALIDATION_COLUMN", "Validation (format)"))
                column_mappings["LOWER_LIMIT_COLUMN"] = st.text_input("Lower Limit Column", column_mappings.get("LOWER_LIMIT_COLUMN", "Lower limit"))
                column_mappings["UPPER_LIMIT_COLUMN"] = st.text_input("Upper Limit Column", column_mappings.get("UPPER_LIMIT_COLUMN", "Upper limit"))

            with col2:
                column_mappings["DEFAULT_VALUE_COLUMN"] = st.text_input("Default Value Column", column_mappings.get("DEFAULT_VALUE_COLUMN", "Default value"))
                column_mappings["CALCULATION_COLUMN"] = st.text_input("Calculation Column", column_mappings.get("CALCULATION_COLUMN", "Calculation"))
                column_mappings["SKIP_LOGIC_COLUMN"] = st.text_input("Skip Logic Column", column_mappings.get("SKIP_LOGIC_COLUMN", "Skip logic"))

            st.subheader("Translation Fields")
            col1, col2 = st.columns(2)
            with col1:
                column_mappings["TRANSLATION_SECTION_COLUMN"] = st.text_input("Translation Section Column", column_mappings.get("TRANSLATION_SECTION_COLUMN", "Translation - Section"))
                column_mappings["TRANSLATION_QUESTION_COLUMN"] = st.text_input("Translation Question Column", column_mappings.get("TRANSLATION_QUESTION_COLUMN", "Translation - Question"))

            with col2:
                column_mappings["TRANSLATION_TOOLTIP_COLUMN"] = st.text_input("Translation Tooltip Column", column_mappings.get("TRANSLATION_TOOLTIP_COLUMN", "Translation - Tooltip"))
                column_mappings["TRANSLATION_ANSWER_COLUMN"] = st.text_input("Translation Answer Column", column_mappings.get("TRANSLATION_ANSWER_COLUMN", "Translation"))

            # Add buttons for saving and resetting
            col1, col2, col3 = st.columns(3)
            with col1:
                save_button = st.form_submit_button("Save Configuration")

            with col2:
                reset_button = st.form_submit_button("Reset to Defaults")

            with col3:
                download_button = st.form_submit_button("Download Config")

    with tab2:
        # Create a form for application settings
        with st.form("app_settings_form"):
            st.subheader("Sheet Filter Settings")

            # Add help text explaining the regex
            st.markdown("""
            The sheet filter prefix is used to identify form sheets in your Excel file.
            By default, it looks for sheets starting with "F" followed by 2 digits (e.g., F01, F02).

            You can customize this using regular expressions:
            - `F\\d{2}` matches F followed by exactly 2 digits (F01, F02, etc.)
            - `F\\d+` matches F followed by any number of digits (F1, F01, F123, etc.)
            - `Form\\d+` matches "Form" followed by digits (Form1, Form2, etc.)
            """)

            app_settings["SHEET_FILTER_PREFIX"] = st.text_input(
                "Sheet Filter Prefix (regex)",
                app_settings.get("SHEET_FILTER_PREFIX", "F\\d{2}"),
                help="Regular expression pattern to identify form sheets in your Excel file"
            )

            # Add buttons for saving and resetting
            col1, col2 = st.columns(2)
            with col1:
                save_settings_button = st.form_submit_button("Save Settings")

            with col2:
                reset_settings_button = st.form_submit_button("Reset Settings to Defaults")

    # Handle column mapping form submission
    if save_button:
        # Save to session state
        st.session_state.column_mappings = column_mappings

        # Save to config.json
        config["columns"] = column_mappings
        if save_config(config):
            st.success("Configuration saved successfully!")

        # Store in localStorage using Streamlit's component
        st.markdown(
            """
            <script>
            localStorage.setItem('formGeneratorConfig', JSON.stringify(
                """ + json.dumps(column_mappings) + """
            ));
            </script>
            """,
            unsafe_allow_html=True
        )

    if reset_button:
        # Reset to defaults
        default_mappings = get_default_column_mappings()
        st.session_state.column_mappings = default_mappings
        config["columns"] = default_mappings
        if save_config(config):
            st.success("Configuration reset to defaults!")
        st.rerun()

    if download_button:
        # Create a downloadable config file
        config_json = json.dumps({"columns": column_mappings, "settings": app_settings}, indent=4)
        b64 = base64.b64encode(config_json.encode()).decode()
        href = f'<a href="data:file/json;base64,{b64}" download="config.json">Download config.json</a>'
        st.markdown(href, unsafe_allow_html=True)

    # Handle application settings form submission
    if save_settings_button:
        # Save to session state
        st.session_state.app_settings = app_settings

        # Save to config.json
        config["settings"] = app_settings
        if save_config(config):
            st.success("Settings saved successfully!")

    if reset_settings_button:
        # Reset to defaults
        default_settings = get_default_app_settings()
        st.session_state.app_settings = default_settings
        config["settings"] = default_settings
        if save_config(config):
            st.success("Settings reset to defaults!")
        st.rerun()

def main():
    # Set page config first - this must be the first Streamlit command
    st.set_page_config(
        page_title="OpenMRS Form Generator",
        page_icon="🏥",
        layout="wide"
    )

    # Initialize session state for JSON preview toggles and generated forms
    if 'initialized' not in st.session_state:
        st.session_state.initialized = True
        st.session_state.generated_forms = []
        st.session_state.temp_file_path = None
        st.session_state.selected_sheets = []
        st.session_state.forms_generated = False
        st.session_state.option_sets_initialized = False
        st.session_state.current_page = "home"

        # Try to load saved config from localStorage
        st.markdown(
            """
            <script>
            const savedConfig = localStorage.getItem('formGeneratorConfig');
            if (savedConfig) {
                window.parent.postMessage({
                    type: 'streamlit:setComponentValue',
                    value: JSON.parse(savedConfig),
                    key: 'loaded_config'
                }, '*');
            }
            </script>
            """,
            unsafe_allow_html=True
        )

    # Create a sidebar container with custom CSS for vertical alignment
    st.sidebar.markdown(
        """
        <style>
        [data-testid="stSidebar"] {
            display: flex;
            flex-direction: column;
        }
        .sidebar-content {
            flex-grow: 1;
        }
        .sidebar-footer {
            margin-top: auto;
            text-align: center;
            padding-bottom: 20px;
        }
        </style>
        <div class="sidebar-content">
        """,
        unsafe_allow_html=True
    )

    # Navigation menu
    st.sidebar.title("Navigation")
    page = st.sidebar.radio(
        "Select page",
        ["O3 Form Generator", "Configuration"],
        key="navigation",
        label_visibility="hidden"
    )

    # Close the content div and start the footer
    st.sidebar.markdown("</div><div class='sidebar-footer'>", unsafe_allow_html=True)

    # Add "Powered by" text and Madiro logo at the bottom of the sidebar
    st.sidebar.markdown("<p style='color: #888; font-size: 0.8em;'>Powered by</p>", unsafe_allow_html=True)
    st.sidebar.image(
        "https://raw.githubusercontent.com/MadiroGlobalHealth/clinical-content-tools/refs/heads/main/.github/workflows/madiro.png",
        width=100
    )

    # Add version number (git commit hash) below the logo
    commit_hash = get_git_commit()
    commit_date_unix = get_git_commit_date()
    try:
        import datetime
        commit_date_str = datetime.datetime.fromtimestamp(int(commit_date_unix), datetime.UTC).strftime('%Y-%m-%d')
    except Exception:
        commit_date_str = "unknown"

    st.sidebar.markdown(f"<p style='color: #888; font-size: 0.7em;'>Version: {commit_hash} - {commit_date_str}</p>", unsafe_allow_html=True)

    # Close the footer div
    st.sidebar.markdown("</div>", unsafe_allow_html=True)

    if page == "O3 Form Generator":
        st.session_state.current_page = "home"
        show_home_page()
    elif page == "Configuration":
        st.session_state.current_page = "config"
        show_configuration_page()

def show_home_page():
    st.title("🏥 OpenMRS 3 Form Generator")
    st.markdown("""
    ### Generate OpenMRS 3 Form Schemas from Excel Metadata

    Verify the mappings with the column names in your Excel file in the Configuration page, then upload an Excel file containing form metadata to generate JSON schemas.
    """)

    # File uploader
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xlsm', 'xltx', 'xltm'],
        help="Upload an Excel file with form metadata. Supported formats: .xlsx, .xlsm, .xltx, .xltm"
    )

    if uploaded_file is not None:
        # Use safe file handler for robust file processing
        with st.spinner('🔄 Processing file... Please wait.'):
            temp_file_path, error_message = safe_file_handler(uploaded_file)
            
            if error_message:
                st.error(f"❌ {error_message}")
                st.stop()
                
            if not temp_file_path:
                st.error("❌ Failed to process uploaded file.")
                st.stop()
                
            st.success(f"✅ File processed successfully: {uploaded_file.name}")
            
            # Store temp file path for cleanup later
            if 'temp_files_to_cleanup' not in st.session_state:
                st.session_state.temp_files_to_cleanup = []
            st.session_state.temp_files_to_cleanup.append(temp_file_path)

        # Initialize option sets with enhanced error handling
        if not st.session_state.option_sets_initialized:
            with st.spinner('🔄 Initializing option sets... Please wait.'):
                try:
                    initialize_option_sets(temp_file_path)
                    st.session_state.option_sets_initialized = True
                    st.success("✅ Option sets initialized successfully")
                except zipfile.BadZipFile:
                    st.error("❌ The uploaded file appears to be corrupted or not a valid Excel file.")
                    st.info("💡 Please try re-saving your Excel file or creating a new one. Make sure to save it in .xlsx format.")
                    cleanup_temp_file(temp_file_path)
                    st.stop()
                except KeyError:
                    st.error("❌ Missing 'OptionSets' sheet in the Excel file.")
                    st.info("💡 Please ensure your Excel file contains a sheet named 'OptionSets' with the expected format.")
                    cleanup_temp_file(temp_file_path)
                    st.stop()
                except MemoryError:
                    st.error("❌ File too large to process - insufficient memory for option sets.")
                    st.info("💡 Try reducing the size of your OptionSets sheet or the overall file size.")
                    cleanup_temp_file(temp_file_path)
                    st.stop()
                except Exception as e:
                    st.error(f"❌ Error initializing option sets: {str(e)}")
                    st.info("💡 Please ensure your Excel file has a sheet named 'OptionSets' with the expected format.")
                    logger.error(f"Option sets initialization failed: {str(e)}")
                    cleanup_temp_file(temp_file_path)
                    st.stop()

        # Update environment variable for metadata file path
        os.environ['METADATA_FILEPATH'] = temp_file_path

        # Extract sheet names with enhanced error handling
        try:
            wb = openpyxl.load_workbook(temp_file_path, read_only=True, data_only=True)
            all_sheet_names = wb.sheetnames
            wb.close()
            logger.info(f"Retrieved {len(all_sheet_names)} sheet names from Excel file")
        except Exception as e:
            st.error(f"❌ Error reading sheet names: {str(e)}")
            cleanup_temp_file(temp_file_path)
            st.stop()

        # Get the configured sheet filter prefix from settings
        config = load_config()
        sheet_filter_prefix = config.get("settings", {}).get("SHEET_FILTER_PREFIX", "F\\d{2}")

        # If filter prefix is empty, show all sheets without filtering
        if not sheet_filter_prefix or sheet_filter_prefix.strip() == "":
            form_sheet_names = all_sheet_names
        else:
            # Filter sheets based on the configured prefix
            form_sheet_names = [sheet for sheet in all_sheet_names if re.match(f'^{sheet_filter_prefix}', sheet)]

            # If no sheets match the filter, show all sheets
            if not form_sheet_names:
                form_sheet_names = all_sheet_names
                st.info("No sheets matched the configured filter. Showing all sheets.")

        # Sheet selection with checkboxes
        st.subheader("Select Sheets to Generate Forms")

        # Create columns for checkboxes to make better use of space
        num_cols = 3  # Number of columns for checkboxes
        cols = st.columns(num_cols)

        selected_sheets = []
        for i, sheet in enumerate(form_sheet_names):
            col_idx = i % num_cols
            with cols[col_idx]:
                if st.checkbox(sheet, key=f"sheet_{sheet}"):
                    selected_sheets.append(sheet)

        # Show selected count
        if selected_sheets:
            st.info(f"Selected {len(selected_sheets)} sheets: {', '.join(selected_sheets)}")

        # Generate forms button
        generate_button = st.button("Generate Forms", type="primary")

        if generate_button or st.session_state.forms_generated:
            if not selected_sheets and not st.session_state.forms_generated:
                st.warning("Please select at least one sheet")
            else:
                # Only generate forms if they haven't been generated yet or if the button was just clicked
                if generate_button and not st.session_state.forms_generated:
                    # Create a full-page spinner overlay
                    with st.spinner('🔄 Generating forms... Please wait, this may take a few minutes.'):
                        st.session_state.temp_file_path = temp_file_path
                        st.session_state.selected_sheets = selected_sheets
                        
                        # Generate forms with progress tracking
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        try:
                            st.session_state.generated_forms = generate_forms_from_sheets(temp_file_path, selected_sheets)
                            progress_bar.progress(100)
                            status_text.text("✅ Forms generated successfully!")
                            st.session_state.forms_generated = True
                            
                        except Exception as e:
                            st.error(f"❌ Error during form generation: {str(e)}")
                            logger.error(f"Form generation failed: {str(e)}")
                            progress_bar.empty()
                            status_text.empty()
                        finally:
                            # Clean up progress indicators
                            import time
                            time.sleep(1)
                            progress_bar.empty()
                            status_text.empty()

                # Display results using the stored generated forms
                st.subheader("Generated Forms")

                for form in st.session_state.generated_forms:
                    st.markdown(f"### Form: {form['sheet']}")

                    # Create a 2x2 grid for metrics
                    metric_cols = st.columns(4)
                    with metric_cols[0]:
                        st.metric("Questions", form['total_questions'])
                    with metric_cols[1]:
                        st.metric("Answers", form['total_answers'])
                    with metric_cols[2]:
                        st.metric("Pages", form.get('num_pages', 'N/A'))
                    with metric_cols[3]:
                        st.metric("Sections", form.get('num_sections', 'N/A'))

                    # Display missing optionSets if any
                    missing_option_sets = form.get('missing_option_sets', [])
                    if missing_option_sets:
                        st.warning(f"⚠️ Found {len(missing_option_sets)} missing optionSets in this form")
                        with st.expander("View missing optionSets"):
                            for missing in missing_option_sets:
                                st.markdown(f"**Question ID:** {missing['question_id']}")
                                st.markdown(f"**Question Label:** {missing['question_label']}")
                                st.markdown(f"**Missing OptionSet:** {missing['optionSet_name']}")
                                st.markdown("---")

                    # Add generation stats
                    stat_cols = st.columns(3)
                    with stat_cols[0]:
                        st.metric("Generation Time", f"{form.get('generation_time', 0):.2f}s")
                    with stat_cols[1]:
                        form_size_kb = form.get('form_size', 0) / 1024
                        st.metric("Form Size", f"{form_size_kb:.1f} KB")
                    with stat_cols[2]:
                        trans_size_kb = form.get('translation_size', 0) / 1024
                        st.metric("Translation Size", f"{trans_size_kb:.1f} KB")

                    # Create columns for form and translation
                    col1, col2 = st.columns(2)

                    with col1:
                        # Form JSON
                        form_json = form['form_json']

                        # Download button
                        st.download_button(
                            label="Download Form JSON",
                            data=form_json,
                            file_name=os.path.basename(form['form_path']),
                            mime='application/json'
                        )

                        # Add collapsible JSON preview
                        with st.expander("Preview Form JSON (click to expand)"):
                            st.code(form_json, language="json")

                    with col2:
                        # Translation JSON
                        translation_json = form['translation_json']

                        # Download button
                        st.download_button(
                            label="Download Translation JSON",
                            data=translation_json,
                            file_name=os.path.basename(form['translation_path']),
                            mime='application/json'
                        )

                        # Add collapsible JSON preview
                        with st.expander("Preview Translation JSON (click to expand)"):
                            st.code(translation_json, language="json")

                    st.markdown("---")  # Add a separator between forms

def cleanup_session_temp_files():
    """Clean up temporary files stored in session state."""
    if 'temp_files_to_cleanup' in st.session_state:
        for temp_file in st.session_state.temp_files_to_cleanup:
            cleanup_temp_file(temp_file)
        st.session_state.temp_files_to_cleanup = []

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"Application error: {str(e)}")
        st.error(f"❌ Application error: {str(e)}")
    finally:
        # Clean up any temporary files on exit
        cleanup_session_temp_files()
